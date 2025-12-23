"""Multi-format document generator supporting PDF, CSV, and Excel."""

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Union
from uuid import uuid4

from src.services.document_generation.template_engine import TemplateEngine
from src.services.document_generation.pdf_service import PDFService, PDFConfig

logger = logging.getLogger(__name__)


class DocumentFormat(str, Enum):
    """Supported document output formats."""
    PDF = "pdf"
    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"
    HTML = "html"


class DocumentStatus(str, Enum):
    """Document generation status."""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"


@dataclass
class DocumentColumn:
    """Column definition for tabular documents."""
    key: str
    header: str
    width: Optional[int] = None
    format_type: Optional[str] = None  # 'currency', 'date', 'number', 'percentage'
    alignment: str = "left"
    visible: bool = True


@dataclass
class DocumentGenerationRequest:
    """Request for document generation."""
    format: DocumentFormat
    template_name: Optional[str] = None
    data: Dict[str, Any] = field(default_factory=dict)
    columns: List[DocumentColumn] = field(default_factory=list)
    title: str = "Document"
    filename: Optional[str] = None
    pdf_config: Optional[PDFConfig] = None
    include_headers: bool = True
    sheet_name: str = "Sheet1"


@dataclass
class DocumentGenerationResult:
    """Result of document generation."""
    id: str
    status: DocumentStatus
    format: DocumentFormat
    filename: str
    content: Optional[bytes] = None
    content_type: str = ""
    size_bytes: int = 0
    error_message: Optional[str] = None
    created_at: datetime = field(default_factory=datetime.utcnow)
    completed_at: Optional[datetime] = None
    progress: int = 0  # 0-100


class DocumentGenerationTask:
    """Represents an ongoing document generation task for progress tracking."""
    
    def __init__(self, task_id: str, format: DocumentFormat, filename: str):
        self.id = task_id
        self.format = format
        self.filename = filename
        self.status = DocumentStatus.PENDING
        self.progress = 0
        self.error_message: Optional[str] = None
        self.created_at = datetime.utcnow()
        self.completed_at: Optional[datetime] = None
        self.content: Optional[bytes] = None
    
    def update_progress(self, progress: int, status: Optional[DocumentStatus] = None):
        """Update task progress."""
        self.progress = min(max(progress, 0), 100)
        if status:
            self.status = status
        if self.progress == 100 and status != DocumentStatus.FAILED:
            self.status = DocumentStatus.COMPLETED
            self.completed_at = datetime.utcnow()
    
    def fail(self, error_message: str):
        """Mark task as failed."""
        self.status = DocumentStatus.FAILED
        self.error_message = error_message
        self.completed_at = datetime.utcnow()
    
    def to_result(self) -> DocumentGenerationResult:
        """Convert to result object."""
        content_types = {
            DocumentFormat.PDF: "application/pdf",
            DocumentFormat.CSV: "text/csv",
            DocumentFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            DocumentFormat.JSON: "application/json",
            DocumentFormat.HTML: "text/html",
        }
        
        return DocumentGenerationResult(
            id=self.id,
            status=self.status,
            format=self.format,
            filename=self.filename,
            content=self.content,
            content_type=content_types.get(self.format, "application/octet-stream"),
            size_bytes=len(self.content) if self.content else 0,
            error_message=self.error_message,
            created_at=self.created_at,
            completed_at=self.completed_at,
            progress=self.progress,
        )


class DocumentGenerator:
    """
    Multi-format document generator.
    
    Supports:
    - PDF generation from templates
    - CSV export with custom columns
    - Excel spreadsheets with formatting
    - JSON data export
    - HTML documents
    
    Features:
    - Progress tracking for large documents
    - Template-based content rendering
    - Custom column definitions
    - Data transformation and formatting
    """
    
    # Content type mappings
    CONTENT_TYPES = {
        DocumentFormat.PDF: "application/pdf",
        DocumentFormat.CSV: "text/csv",
        DocumentFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        DocumentFormat.JSON: "application/json",
        DocumentFormat.HTML: "text/html",
    }
    
    # File extensions
    FILE_EXTENSIONS = {
        DocumentFormat.PDF: ".pdf",
        DocumentFormat.CSV: ".csv",
        DocumentFormat.EXCEL: ".xlsx",
        DocumentFormat.JSON: ".json",
        DocumentFormat.HTML: ".html",
    }
    
    def __init__(
        self,
        template_engine: Optional[TemplateEngine] = None,
        pdf_service: Optional[PDFService] = None
    ):
        """
        Initialize the document generator.
        
        Args:
            template_engine: Template engine for content rendering
            pdf_service: PDF generation service
        """
        self.template_engine = template_engine or TemplateEngine()
        self.pdf_service = pdf_service or PDFService(self.template_engine)
        self._tasks: Dict[str, DocumentGenerationTask] = {}
    
    def generate(self, request: DocumentGenerationRequest) -> DocumentGenerationResult:
        """
        Generate a document synchronously.
        
        Args:
            request: Document generation request
            
        Returns:
            Document generation result with content
        """
        task_id = str(uuid4())
        filename = self._get_filename(request)
        task = DocumentGenerationTask(task_id, request.format, filename)
        self._tasks[task_id] = task
        
        try:
            task.status = DocumentStatus.PROCESSING
            task.update_progress(10)
            
            if request.format == DocumentFormat.PDF:
                content = self._generate_pdf(request, task)
            elif request.format == DocumentFormat.CSV:
                content = self._generate_csv(request, task)
            elif request.format == DocumentFormat.EXCEL:
                content = self._generate_excel(request, task)
            elif request.format == DocumentFormat.JSON:
                content = self._generate_json(request, task)
            elif request.format == DocumentFormat.HTML:
                content = self._generate_html(request, task)
            else:
                raise ValueError(f"Unsupported format: {request.format}")
            
            task.content = content
            task.update_progress(100, DocumentStatus.COMPLETED)
            
        except Exception as e:
            logger.exception(f"Document generation failed: {e}")
            task.fail(str(e))
        
        return task.to_result()
    
    async def generate_async(
        self,
        request: DocumentGenerationRequest,
        progress_callback: Optional[callable] = None
    ) -> str:
        """
        Start asynchronous document generation.
        
        Args:
            request: Document generation request
            progress_callback: Optional callback for progress updates
            
        Returns:
            Task ID for tracking progress
        """
        task_id = str(uuid4())
        filename = self._get_filename(request)
        task = DocumentGenerationTask(task_id, request.format, filename)
        self._tasks[task_id] = task
        
        # In a real implementation, this would dispatch to a background worker
        # For now, we'll generate synchronously
        try:
            task.status = DocumentStatus.PROCESSING
            if progress_callback:
                await progress_callback(task_id, 10, DocumentStatus.PROCESSING)
            
            result = self.generate(request)
            task.content = result.content
            task.status = result.status
            task.error_message = result.error_message
            
            if progress_callback:
                await progress_callback(task_id, 100, task.status)
                
        except Exception as e:
            task.fail(str(e))
            if progress_callback:
                await progress_callback(task_id, 0, DocumentStatus.FAILED)
        
        return task_id
    
    def get_task_status(self, task_id: str) -> Optional[DocumentGenerationResult]:
        """
        Get the status of a document generation task.
        
        Args:
            task_id: Task identifier
            
        Returns:
            Task result or None if not found
        """
        task = self._tasks.get(task_id)
        if task:
            return task.to_result()
        return None
    
    def get_task_content(self, task_id: str) -> Optional[bytes]:
        """
        Get the generated document content.
        
        Args:
            task_id: Task identifier
            
        Returns:
            Document content bytes or None
        """
        task = self._tasks.get(task_id)
        if task and task.status == DocumentStatus.COMPLETED:
            return task.content
        return None
    
    def cancel_task(self, task_id: str) -> bool:
        """
        Cancel a document generation task.
        
        Args:
            task_id: Task identifier
            
        Returns:
            True if cancelled, False if not found or already completed
        """
        task = self._tasks.get(task_id)
        if task and task.status in (DocumentStatus.PENDING, DocumentStatus.PROCESSING):
            task.status = DocumentStatus.CANCELLED
            task.completed_at = datetime.utcnow()
            return True
        return False
    
    def _get_filename(self, request: DocumentGenerationRequest) -> str:
        """Generate filename for the document."""
        if request.filename:
            base_name = request.filename
        else:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            safe_title = "".join(c for c in request.title if c.isalnum() or c in " -_")
            base_name = f"{safe_title}_{timestamp}"
        
        ext = self.FILE_EXTENSIONS.get(request.format, "")
        if not base_name.endswith(ext):
            base_name += ext
        
        return base_name
    
    def _generate_pdf(
        self,
        request: DocumentGenerationRequest,
        task: DocumentGenerationTask
    ) -> bytes:
        """Generate PDF document."""
        task.update_progress(20)
        
        if request.template_name:
            content = self.pdf_service.generate_from_template(
                request.template_name,
                request.data,
                request.pdf_config
            )
        else:
            # Generate from data
            html = self._data_to_html(request)
            config = request.pdf_config or PDFConfig()
            config.metadata.title = request.title
            content = self.pdf_service.generate(html, config, request.data)
        
        task.update_progress(90)
        return content
    
    def _generate_csv(
        self,
        request: DocumentGenerationRequest,
        task: DocumentGenerationTask
    ) -> bytes:
        """Generate CSV document."""
        task.update_progress(20)
        
        output = io.StringIO()
        rows = request.data.get("rows", [])
        columns = request.columns or self._infer_columns(rows)
        
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        
        # Write headers
        if request.include_headers:
            headers = [col.header for col in columns if col.visible]
            writer.writerow(headers)
        
        task.update_progress(40)
        
        # Write data rows
        total_rows = len(rows)
        for idx, row in enumerate(rows):
            csv_row = []
            for col in columns:
                if not col.visible:
                    continue
                value = self._get_nested_value(row, col.key)
                formatted = self._format_value(value, col.format_type)
                csv_row.append(formatted)
            writer.writerow(csv_row)
            
            # Update progress
            if total_rows > 0:
                progress = 40 + int((idx + 1) / total_rows * 50)
                task.update_progress(progress)
        
        task.update_progress(95)
        return output.getvalue().encode('utf-8')
    
    def _generate_excel(
        self,
        request: DocumentGenerationRequest,
        task: DocumentGenerationTask
    ) -> bytes:
        """Generate Excel document."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV")
            return self._generate_csv(request, task)
        
        task.update_progress(20)
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = request.sheet_name[:31]  # Excel limit
        
        rows = request.data.get("rows", [])
        columns = request.columns or self._infer_columns(rows)
        visible_columns = [col for col in columns if col.visible]
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Write headers
        if request.include_headers:
            for col_idx, col in enumerate(visible_columns, 1):
                cell = sheet.cell(row=1, column=col_idx, value=col.header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # Set column width
                if col.width:
                    sheet.column_dimensions[get_column_letter(col_idx)].width = col.width
                else:
                    sheet.column_dimensions[get_column_letter(col_idx)].width = max(len(col.header) + 2, 12)
        
        task.update_progress(40)
        
        # Write data rows
        start_row = 2 if request.include_headers else 1
        total_rows = len(rows)
        
        for row_idx, row_data in enumerate(rows):
            for col_idx, col in enumerate(visible_columns, 1):
                value = self._get_nested_value(row_data, col.key)
                formatted = self._format_value_for_excel(value, col.format_type)
                
                cell = sheet.cell(row=start_row + row_idx, column=col_idx, value=formatted)
                cell.border = thin_border
                
                # Apply alignment
                if col.alignment == "right":
                    cell.alignment = Alignment(horizontal="right")
                elif col.alignment == "center":
                    cell.alignment = Alignment(horizontal="center")
                
                # Apply number formats
                if col.format_type == "currency":
                    cell.number_format = '"$"#,##0.00'
                elif col.format_type == "percentage":
                    cell.number_format = '0.00%'
                elif col.format_type == "date":
                    cell.number_format = 'YYYY-MM-DD'
            
            # Update progress
            if total_rows > 0:
                progress = 40 + int((row_idx + 1) / total_rows * 50)
                task.update_progress(progress)
        
        task.update_progress(95)
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()
    
    def _generate_json(
        self,
        request: DocumentGenerationRequest,
        task: DocumentGenerationTask
    ) -> bytes:
        """Generate JSON document."""
        task.update_progress(50)
        
        output = {
            "title": request.title,
            "generated_at": datetime.utcnow().isoformat(),
            "data": request.data
        }
        
        task.update_progress(90)
        return json.dumps(output, indent=2, default=str).encode('utf-8')
    
    def _generate_html(
        self,
        request: DocumentGenerationRequest,
        task: DocumentGenerationTask
    ) -> bytes:
        """Generate HTML document."""
        task.update_progress(20)
        
        if request.template_name and request.template_name in self.template_engine.partials:
            html = self.template_engine.render(
                self.template_engine.partials[request.template_name],
                request.data
            )
        else:
            html = self._data_to_html(request)
        
        # Wrap in full HTML document
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{request.title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4472C4; color: white; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>{request.title}</h1>
            <p>Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
            {html}
        </body>
        </html>
        """
        
        task.update_progress(90)
        return full_html.encode('utf-8')
    
    def _data_to_html(self, request: DocumentGenerationRequest) -> str:
        """Convert data to HTML table."""
        rows = request.data.get("rows", [])
        if not rows:
            return "<p>No data available.</p>"
        
        columns = request.columns or self._infer_columns(rows)
        visible_columns = [col for col in columns if col.visible]
        
        html_parts = ["<table>"]
        
        # Headers
        if request.include_headers:
            html_parts.append("<thead><tr>")
            for col in visible_columns:
                html_parts.append(f"<th>{col.header}</th>")
            html_parts.append("</tr></thead>")
        
        # Body
        html_parts.append("<tbody>")
        for row in rows:
            html_parts.append("<tr>")
            for col in visible_columns:
                value = self._get_nested_value(row, col.key)
                formatted = self._format_value(value, col.format_type)
                html_parts.append(f"<td>{formatted}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody></table>")
        
        return "".join(html_parts)
    
    def _infer_columns(self, rows: List[Dict[str, Any]]) -> List[DocumentColumn]:
        """Infer column definitions from data."""
        if not rows:
            return []
        
        columns = []
        first_row = rows[0]
        
        def add_columns(data: Dict[str, Any], prefix: str = ""):
            for key, value in data.items():
                full_key = f"{prefix}.{key}" if prefix else key
                if isinstance(value, dict):
                    add_columns(value, full_key)
                else:
                    header = key.replace("_", " ").title()
                    columns.append(DocumentColumn(key=full_key, header=header))
        
        add_columns(first_row)
        return columns
    
    def _get_nested_value(self, data: Dict[str, Any], key: str) -> Any:
        """Get a nested value from a dictionary using dot notation."""
        keys = key.split(".")
        value = data
        
        for k in keys:
            if isinstance(value, dict):
                value = value.get(k)
            elif hasattr(value, k):
                value = getattr(value, k)
            else:
                return None
            
            if value is None:
                return None
        
        return value
    
    def _format_value(self, value: Any, format_type: Optional[str]) -> str:
        """Format a value based on type."""
        if value is None:
            return ""
        
        if format_type == "currency":
            try:
                return f"${float(value):,.2f}"
            except (ValueError, TypeError):
                return str(value)
        elif format_type == "date":
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            return str(value)
        elif format_type == "number":
            try:
                return f"{float(value):,.0f}"
            except (ValueError, TypeError):
                return str(value)
        elif format_type == "percentage":
            try:
                return f"{float(value) * 100:.1f}%"
            except (ValueError, TypeError):
                return str(value)
        
        return str(value)
    
    def _format_value_for_excel(self, value: Any, format_type: Optional[str]) -> Any:
        """Format a value for Excel (preserving type when possible)."""
        if value is None:
            return ""
        
        if format_type in ("currency", "number", "percentage"):
            try:
                return float(value)
            except (ValueError, TypeError):
                return value
        elif format_type == "date":
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                try:
                    return datetime.fromisoformat(value.replace("Z", "+00:00"))
                except ValueError:
                    pass
        
        return value

